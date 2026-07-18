"""
云汐 M9 数据水晶 - Excel 连接器

P3 优化：数据采集管道 + 连接器生态
Excel 文件连接器，支持多 Sheet、表头配置、样式保留
"""

from __future__ import annotations

import logging
from typing import Iterator, List, Dict, Any, Optional
from pathlib import Path

from .base import (
    BaseConnector,
    ConnectorMeta,
    ConnectorRegistry,
    ConnectorType,
    ConnectionStatus,
)

logger = logging.getLogger(__name__)


@ConnectorRegistry.register
class ExcelConnector(BaseConnector):
    """
    Excel 文件连接器

    特性：
    - 多 Sheet 支持
    - 表头配置（自定义表头行）
    - 样式保留（写入时）
    - 批量读取/写入
    - 支持 xlsx / xls

    依赖：openpyxl（xlsx）/ xlrd（xls，只读）
    """

    meta = ConnectorMeta(
        name="excel",
        connector_type=ConnectorType.FILE,
        description="Excel 文件连接器，支持多 Sheet、自定义表头",
        version="1.0.0",
        supported_operations=["read", "write", "batch_read", "batch_write", "list_tables", "schema"],
    )

    def __init__(self, config: Optional[Dict[str, Any]] = None):
        super().__init__(config)
        self._file_path: str = ""
        self._sheet_name: Optional[str] = None
        self._header_row: int = 1  # 表头所在行（1-based）
        self._start_row: int = 2   # 数据起始行（1-based）
        self._workbook = None
        self._worksheet = None

    def connect(self, config: Optional[Dict[str, Any]] = None) -> bool:
        """
        连接（打开）Excel 文件

        config 参数：
        - file_path: Excel 文件路径
        - sheet_name: 默认 Sheet 名称（可选，默认第一个 Sheet）
        - header_row: 表头行号（1-based，默认 1）
        - start_row: 数据起始行（1-based，默认 header_row + 1）
        - mode: 打开模式（r/w，默认 r）
        """
        if config:
            self._config.update(config)

        self._status = ConnectionStatus.CONNECTING
        try:
            try:
                import openpyxl
            except ImportError:
                self._status = ConnectionStatus.ERROR
                self._last_error = "openpyxl 未安装，请执行 pip install openpyxl"
                logger.warning(self._last_error)
                return False

            self._file_path = self._config.get("file_path", "")
            if not self._file_path:
                raise ValueError("必须指定 file_path")

            self._sheet_name = self._config.get("sheet_name")
            self._header_row = int(self._config.get("header_row", 1))
            self._start_row = int(self._config.get("start_row", self._header_row + 1))
            mode = self._config.get("mode", "r")

            # 确保目录存在
            Path(self._file_path).parent.mkdir(parents=True, exist_ok=True)

            # 打开工作簿
            if mode == "r" and Path(self._file_path).exists():
                self._workbook = openpyxl.load_workbook(self._file_path, read_only=True, data_only=True)
            elif mode == "w" or not Path(self._file_path).exists():
                self._workbook = openpyxl.Workbook()
            else:
                self._workbook = openpyxl.load_workbook(self._file_path)

            # 选择工作表
            if self._sheet_name:
                if self._sheet_name in self._workbook.sheetnames:
                    self._worksheet = self._workbook[self._sheet_name]
                else:
                    # 创建新 sheet
                    self._worksheet = self._workbook.create_sheet(self._sheet_name)
            else:
                if self._workbook.sheetnames:
                    self._worksheet = self._workbook[self._workbook.sheetnames[0]]
                else:
                    self._worksheet = self._workbook.active

            self._status = ConnectionStatus.CONNECTED
            self._stats.connection_count += 1
            logger.info(f"Excel 文件已打开: {self._file_path}")
            return True

        except Exception as e:
            self._status = ConnectionStatus.ERROR
            self._last_error = str(e)
            self._record_error()
            logger.error(f"Excel 连接失败: {e}")
            return False

    def disconnect(self) -> bool:
        """关闭 Excel 文件"""
        try:
            if self._workbook:
                mode = self._config.get("mode", "r")
                if mode == "w" or mode == "a":
                    self._workbook.save(self._file_path)
                self._workbook.close()
                self._workbook = None
                self._worksheet = None
            self._status = ConnectionStatus.DISCONNECTED
            logger.info("Excel 文件已关闭")
            return True
        except Exception as e:
            self._last_error = str(e)
            self._record_error()
            return False

    def _get_headers(self, worksheet=None) -> List[str]:
        """获取表头"""
        ws = worksheet or self._worksheet
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=self._header_row, column=col).value
            headers.append(str(cell_value) if cell_value is not None else f"column_{col}")
        return headers

    def read(self, query: Optional[Dict[str, Any]] = None) -> Iterator[Dict[str, Any]]:
        """
        流式读取 Excel 数据

        query 参数：
        - sheet_name: Sheet 名称（可选）
        - limit: 最大读取行数
        - offset: 跳过行数
        - columns: 只读取指定列
        - filter: 过滤条件 dict
        """
        self._ensure_connected()
        query = query or {}

        try:
            import openpyxl

            sheet_name = query.get("sheet_name", self._sheet_name)
            limit = query.get("limit")
            offset = query.get("offset", 0)
            columns = query.get("columns")
            filter_cond = query.get("filter", {})

            # 选择工作表
            ws = self._worksheet
            if sheet_name and sheet_name in self._workbook.sheetnames:
                ws = self._workbook[sheet_name]

            headers = self._get_headers(ws)

            count = 0
            skipped = 0

            for row_idx in range(self._start_row, ws.max_row + 1):
                # 跳过
                if skipped < offset:
                    skipped += 1
                    continue

                # 读取行数据
                row_data = {}
                for col_idx, header in enumerate(headers, start=1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    row_data[header] = cell_value

                # 检查是否为空行
                if all(v is None or v == "" for v in row_data.values()):
                    continue

                # 过滤
                if filter_cond:
                    match = True
                    for key, value in filter_cond.items():
                        if str(row_data.get(key, "")) != str(value):
                            match = False
                            break
                    if not match:
                        continue

                # 只保留指定列
                if columns:
                    row_data = {col: row_data.get(col) for col in columns}

                count += 1
                yield row_data

                if limit and count >= limit:
                    break

            self._record_read(count=count, bytes_read=count * 100)

        except Exception as e:
            self._record_error()
            logger.error(f"Excel 读取失败: {e}")
            raise

    def read_batch(self, batch_size: int = 100, query: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
        """批量读取"""
        query = query or {}
        query["limit"] = batch_size
        return super().read_batch(batch_size, query)

    def write(self, data: List[Dict[str, Any]]) -> int:
        """
        批量写入 Excel

        配置：
        - sheet_name: 目标 Sheet 名称
        - write_mode: overwrite / append（默认 append）
        - include_header: 是否写入表头（默认 True）
        - preserve_style: 是否保留样式（默认 False）
        """
        self._ensure_connected()

        if not data:
            return 0

        try:
            import openpyxl

            # 确定目标 Sheet：优先 write_sheet，其次 _sheet_name，最后用当前活动 sheet
            sheet_name = self._config.get("write_sheet")
            if not sheet_name:
                sheet_name = self._sheet_name
            if not sheet_name and self._worksheet:
                sheet_name = self._worksheet.title
            if not sheet_name:
                sheet_name = "Sheet1"
            write_mode = self._config.get("write_mode", "append")
            include_header = self._config.get("include_header", True)

            # 获取或创建工作表
            if sheet_name in self._workbook.sheetnames:
                ws = self._workbook[sheet_name]
                if write_mode == "overwrite":
                    # 清空工作表
                    self._workbook.remove(ws)
                    ws = self._workbook.create_sheet(sheet_name)
            else:
                ws = self._workbook.create_sheet(sheet_name)

            columns = list(data[0].keys())
            start_row = 1

            # 写入表头
            if include_header and (ws.max_row == 1 and ws.cell(1, 1).value is None):
                for col_idx, col_name in enumerate(columns, start=1):
                    ws.cell(row=1, column=col_idx, value=col_name)
                start_row = 2
            elif include_header and write_mode == "overwrite":
                for col_idx, col_name in enumerate(columns, start=1):
                    ws.cell(row=1, column=col_idx, value=col_name)
                start_row = 2
            else:
                # 追加模式：找到最后一行
                start_row = ws.max_row + 1

            # 写入数据
            for row_offset, record in enumerate(data):
                for col_idx, col_name in enumerate(columns, start=1):
                    ws.cell(row=start_row + row_offset, column=col_idx, value=record.get(col_name))

            # 保存
            self._workbook.save(self._file_path)

            count = len(data)
            self._record_write(count=count, bytes_written=count * 100)
            return count

        except Exception as e:
            self._record_error()
            logger.error(f"Excel 写入失败: {e}")
            raise

    def list_tables(self) -> List[str]:
        """列出所有 Sheet 名称"""
        self._ensure_connected()
        try:
            return list(self._workbook.sheetnames)
        except Exception as e:
            self._record_error()
            raise

    def get_schema(self, table: str) -> Dict[str, Any]:
        """获取指定 Sheet 的 Schema"""
        self._ensure_connected()
        try:
            if table not in self._workbook.sheetnames:
                raise ValueError(f"Sheet 不存在: {table}")

            ws = self._workbook[table]
            headers = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=self._header_row, column=col).value
                headers.append(str(cell_value) if cell_value is not None else f"column_{col}")

            # 采样几行推断类型
            sample_rows = min(10, ws.max_row - self._start_row + 1)
            type_samples = {h: set() for h in headers}

            for row_idx in range(self._start_row, self._start_row + sample_rows):
                if row_idx > ws.max_row:
                    break
                for col_idx, header in enumerate(headers, start=1):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if val is not None:
                        type_samples[header].add(type(val).__name__)

            fields = {}
            for header in headers:
                types = type_samples[header]
                if len(types) == 1:
                    field_type = list(types)[0]
                elif len(types) > 1:
                    field_type = "mixed"
                else:
                    field_type = "string"

                fields[header] = {
                    "type": field_type,
                    "nullable": len(types) < sample_rows,
                }

            return {
                "table": table,
                "fields": fields,
                "row_count": max(0, ws.max_row - self._start_row + 1),
                "column_count": len(headers),
            }
        except Exception as e:
            self._record_error()
            raise

    def _health_probe(self) -> None:
        """健康探针"""
        if self._workbook:
            _ = self._workbook.sheetnames
